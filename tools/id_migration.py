#!/usr/bin/env python3
"""Export, apply, and audit the approved ERYDON ID migration map.

The committed TSV is the reproducible bridge between the reviewed workbook and
the runtime registry aliases.  Source application is intentionally narrow:
only registry-keyed blockstate/item-model filenames and language keys move.
Internal block-model and texture paths stay stable for resource-pack and CTM
compatibility.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


EXPECTED_TOTAL = 1775
EXPECTED_ALIASES = 1586
EXPECTED_DIRECT_RENAMES = 189
EXPECTED_QUATREFOIL_ALIASES = 48
MANIFEST_COLUMNS = (
    "source_row",
    "mode",
    "old_path",
    "canonical_path",
    "old_display_name",
    "canonical_display_name",
    "design_style",
    "publication_status",
    "reason",
    "review_status",
)
JSON_LINE = re.compile(
    r'^(?P<indent>\s*)"(?P<key>(?:\\.|[^"\\])+)"\s*:\s*'
    r'"(?P<value>(?:\\.|[^"\\])*)"(?P<comma>,?)\s*$'
)


@dataclass(frozen=True)
class Entry:
    source_row: int
    mode: str
    old_path: str
    canonical_path: str
    old_display_name: str
    canonical_display_name: str
    design_style: str
    publication_status: str
    reason: str
    review_status: str

    @property
    def permanent_alias(self) -> bool:
        return self.mode == "PERMANENT_ALIAS"


def parser() -> argparse.ArgumentParser:
    result = argparse.ArgumentParser(description=__doc__)
    subcommands = result.add_subparsers(dest="command", required=True)

    export = subcommands.add_parser("export", help="export approved ERYDON rows from the workbook")
    export.add_argument("--workbook", type=Path, required=True)
    export.add_argument("--output", type=Path, required=True)
    export.add_argument("--sheet", default="05 ID Migration Map")

    apply = subcommands.add_parser("apply", help="apply registry-keyed source moves")
    apply.add_argument("--manifest", type=Path, required=True)
    apply.add_argument("--repo", type=Path, default=Path.cwd())

    audit = subcommands.add_parser("audit", help="audit the committed source migration")
    audit.add_argument("--manifest", type=Path, required=True)
    audit.add_argument("--repo", type=Path, default=Path.cwd())
    audit.add_argument("--report", type=Path)
    return result


def require(condition: bool, message: str) -> None:
    if not condition:
        raise ValueError(message)


def validate_entries(entries: list[Entry]) -> None:
    require(len(entries) == EXPECTED_TOTAL,
            f"expected {EXPECTED_TOTAL} ERYDON migrations, found {len(entries)}")
    aliases = [entry for entry in entries if entry.permanent_alias]
    direct = [entry for entry in entries if entry.mode == "DIRECT_RENAME"]
    require(len(aliases) == EXPECTED_ALIASES,
            f"expected {EXPECTED_ALIASES} permanent aliases, found {len(aliases)}")
    require(len(direct) == EXPECTED_DIRECT_RENAMES,
            f"expected {EXPECTED_DIRECT_RENAMES} direct renames, found {len(direct)}")
    require(all(entry.review_status == "Approved" for entry in entries),
            "every migration row must be Approved")

    old_paths = [entry.old_path for entry in entries]
    canonical_paths = [entry.canonical_path for entry in entries]
    require(len(set(old_paths)) == len(old_paths), "duplicate old paths in migration map")
    require(len(set(canonical_paths)) == len(canonical_paths),
            "duplicate canonical paths in migration map")
    chains = set(old_paths).intersection(canonical_paths)
    require(not chains, f"alias chains are forbidden: {sorted(chains)[:5]}")
    require(all(entry.old_path != entry.canonical_path for entry in entries),
            "migration rows must change the ID")

    quatrefoil = [entry for entry in aliases if "quatrefoil" in entry.old_path]
    require(len(quatrefoil) == EXPECTED_QUATREFOIL_ALIASES,
            f"expected {EXPECTED_QUATREFOIL_ALIASES} Quatrefoil aliases, found {len(quatrefoil)}")


def workbook_entries(workbook: Path, sheet_name: str) -> list[Entry]:
    try:
        from openpyxl import load_workbook
    except ImportError as exception:
        raise RuntimeError("openpyxl is required only for workbook export") from exception

    worksheet = load_workbook(workbook, read_only=True, data_only=True)[sheet_name]
    headers = list(next(worksheet.iter_rows(min_row=4, max_row=4, values_only=True)))
    index = {str(value): position for position, value in enumerate(headers)}
    required = {
        "Mod", "Design style", "Current block ID", "Proposed block ID",
        "Current item ID", "Proposed item ID", "Current display name",
        "Proposed display name", "Publication status", "Migration action",
        "Block alias", "Item alias", "Reason", "Collision check", "Review status",
    }
    missing = required.difference(index)
    require(not missing, f"missing workbook columns: {sorted(missing)}")

    entries: list[Entry] = []
    for source_row, values in enumerate(
            worksheet.iter_rows(min_row=5, values_only=True), start=5):
        if str(values[index["Mod"]] or "").strip().upper() != "ERYDON":
            continue

        action = str(values[index["Migration action"]] or "").strip().upper()
        mode = action.replace(" ", "_")
        require(mode in {"PERMANENT_ALIAS", "DIRECT_RENAME"},
                f"row {source_row}: unsupported migration action {action!r}")

        old_block = str(values[index["Current block ID"]] or "").strip()
        new_block = str(values[index["Proposed block ID"]] or "").strip()
        old_item = str(values[index["Current item ID"]] or "").strip()
        new_item = str(values[index["Proposed item ID"]] or "").strip()
        require(old_block == old_item, f"row {source_row}: block/item old IDs differ")
        require(new_block == new_item, f"row {source_row}: block/item canonical IDs differ")
        require(old_block.startswith("erydon:") and new_block.startswith("erydon:"),
                f"row {source_row}: IDs must stay in the erydon namespace")
        require(str(values[index["Collision check"]] or "").strip() == "Clear",
                f"row {source_row}: collision check is not Clear")

        block_alias = str(values[index["Block alias"]] or "").strip()
        item_alias = str(values[index["Item alias"]] or "").strip()
        expected_alias = "Yes" if mode == "PERMANENT_ALIAS" else "No"
        require(block_alias == expected_alias and item_alias == expected_alias,
                f"row {source_row}: alias flags do not match {mode}")

        entries.append(Entry(
            source_row=source_row,
            mode=mode,
            old_path=old_block.split(":", 1)[1],
            canonical_path=new_block.split(":", 1)[1],
            old_display_name=str(values[index["Current display name"]] or "").strip(),
            canonical_display_name=str(values[index["Proposed display name"]] or "").strip(),
            design_style=str(values[index["Design style"]] or "").strip(),
            publication_status=str(values[index["Publication status"]] or "").strip(),
            reason=str(values[index["Reason"]] or "").strip(),
            review_status=str(values[index["Review status"]] or "").strip(),
        ))

    validate_entries(entries)
    return entries


def write_manifest(entries: Iterable[Entry], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=MANIFEST_COLUMNS, delimiter="\t",
                                lineterminator="\n")
        writer.writeheader()
        for entry in entries:
            writer.writerow(entry.__dict__)


def read_manifest(path: Path) -> list[Entry]:
    with path.open("r", encoding="utf-8", newline="") as stream:
        reader = csv.DictReader(stream, delimiter="\t")
        require(tuple(reader.fieldnames or ()) == MANIFEST_COLUMNS,
                f"unexpected manifest columns in {path}")
        entries = [Entry(source_row=int(row["source_row"]),
                         **{key: row[key] for key in MANIFEST_COLUMNS if key != "source_row"})
                   for row in reader]
    validate_entries(entries)
    return entries


def move_registry_resources(repo: Path, entries: list[Entry]) -> int:
    roots = (
        repo / "src/main/resources/assets/erydon/blockstates",
        repo / "src/main/resources/assets/erydon/models/item",
    )
    moved = 0
    for entry in entries:
        for root in roots:
            old = root / f"{entry.old_path}.json"
            canonical = root / f"{entry.canonical_path}.json"
            require(not (old.exists() and canonical.exists()),
                    f"both old and canonical resources exist: {old} / {canonical}")
            if old.exists():
                old.rename(canonical)
                moved += 1
            else:
                require(canonical.is_file(),
                        f"missing old and canonical registry resource for row {entry.source_row}: {root}")
    return moved


def update_registry_references(repo: Path, entries: list[Entry]) -> int:
    replacements = {
        f"erydon:{entry.old_path}": f"erydon:{entry.canonical_path}"
        for entry in entries
    }
    pattern = re.compile(
        r"(?:"
        + "|".join(re.escape(value) for value in sorted(replacements, key=len, reverse=True))
        + r")(?![a-z0-9_])"
    )
    changed = 0
    registry_reference_files = [
        *sorted((repo / "src/main/resources/data").rglob("*.json")),
        repo / "src/main/resources/assets/erydon/shaders/block.properties",
    ]
    for path in registry_reference_files:
        raw = path.read_text(encoding="utf-8-sig")
        updated, count = pattern.subn(lambda match: replacements[match.group(0)], raw)
        if count:
            path.write_text(updated, encoding="utf-8", newline="")
            changed += count
    return changed


def localized_value(language: str, entry: Entry, current: str) -> str:
    if language == "en_us":
        return entry.canonical_display_name

    value = current
    if "Byzantine style" in entry.reason:
        if language == "de_de":
            adjective = "byzantinisches" if "Gesims" in value else "byzantinische"
            value = re.sub(
                r"guilloche|byzantinisch(?:es|e)?s*",
                adjective,
                value,
                flags=re.IGNORECASE,
            )
        elif language == "es_es":
            adjective = "bizantina" if "Cornisa" in value else "bizantino"
            value = re.sub("guilloche|bizantino|bizantina", adjective, value, flags=re.IGNORECASE)
    if "Rose → Rosette" in entry.reason:
        if language == "de_de":
            value = re.sub(r"\bRose\b", "Rosette", value, flags=re.IGNORECASE)
        elif language == "es_es":
            value = re.sub(r"\bRosa\b", "Rosetón", value, flags=re.IGNORECASE)
    if "Inlay grammar" in entry.reason:
        if language == "de_de":
            value = re.sub(r"\bBlock\b", "Intarsie", value)
        elif language == "es_es":
            value = re.sub(r"\bBloque\b", "Incrustación", value, flags=re.IGNORECASE)
    return value


def update_language(path: Path, language: str, entries: list[Entry]) -> tuple[int, int]:
    raw = path.read_text(encoding="utf-8-sig")
    newline = "\r\n" if "\r\n" in raw else "\n"
    lines = raw.splitlines()
    key_to_entry = {f"block.erydon.{entry.old_path}": entry for entry in entries}
    key_to_entry.update({f"block.erydon.{entry.canonical_path}": entry for entry in entries})
    canonical_seen: set[str] = set()
    changed_keys = 0
    changed_values = 0
    output: list[str] = []

    for line in lines:
        match = JSON_LINE.match(line)
        if match is None or match.group("key") not in key_to_entry:
            output.append(line)
            continue

        entry = key_to_entry[match.group("key")]
        canonical_key = f"block.erydon.{entry.canonical_path}"
        if canonical_key in canonical_seen:
            # Earlier migration work deliberately staged canonical keys beside
            # the published keys.  Collapse that pair into the first canonical
            # entry now that the registry migration is being activated.
            changed_keys += 1
            continue
        canonical_seen.add(canonical_key)

        current_value = json.loads(f'"{match.group("value")}"')
        next_value = localized_value(language, entry, current_value)
        if match.group("key") != canonical_key:
            changed_keys += 1
        if current_value != next_value:
            changed_values += 1
        encoded_value = json.dumps(next_value, ensure_ascii=False)
        output.append(
            f'{match.group("indent")}{json.dumps(canonical_key, ensure_ascii=False)}: '
            f'{encoded_value}{match.group("comma")}'
        )

    missing = {f"block.erydon.{entry.canonical_path}" for entry in entries}.difference(canonical_seen)
    require(not missing, f"missing canonical language keys in {path}: {sorted(missing)[:5]}")
    path.write_text(newline.join(output) + newline, encoding="utf-8", newline="")
    return changed_keys, changed_values


def apply_source(repo: Path, entries: list[Entry]) -> dict[str, int]:
    repo = repo.resolve()
    moved = move_registry_resources(repo, entries)
    changed_registry_references = update_registry_references(repo, entries)
    changed_keys = 0
    changed_values = 0
    for language in ("en_us", "de_de", "es_es"):
        path = repo / f"src/main/resources/assets/erydon/lang/{language}.json"
        keys, values = update_language(path, language, entries)
        changed_keys += keys
        changed_values += values
    return {
        "moved_registry_resources": moved,
        "changed_registry_references": changed_registry_references,
        "changed_language_keys": changed_keys,
        "changed_language_values": changed_values,
    }


def audit_source(repo: Path, entries: list[Entry]) -> dict[str, object]:
    repo = repo.resolve()
    errors: list[str] = []
    canonical_resources = 0
    stale_resources = 0
    for entry in entries:
        for relative_root in (
                "src/main/resources/assets/erydon/blockstates",
                "src/main/resources/assets/erydon/models/item"):
            root = repo / relative_root
            old = root / f"{entry.old_path}.json"
            canonical = root / f"{entry.canonical_path}.json"
            if canonical.is_file():
                canonical_resources += 1
            else:
                errors.append(f"missing canonical resource: {canonical}")
            if old.exists():
                stale_resources += 1
                errors.append(f"stale old registry resource: {old}")

    language_keys = 0
    stale_language_keys = 0
    for language in ("en_us", "de_de", "es_es"):
        path = repo / f"src/main/resources/assets/erydon/lang/{language}.json"
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        for entry in entries:
            canonical_key = f"block.erydon.{entry.canonical_path}"
            old_key = f"block.erydon.{entry.old_path}"
            if canonical_key in data:
                language_keys += 1
            else:
                errors.append(f"missing language key in {language}: {canonical_key}")
            if old_key in data:
                stale_language_keys += 1
                errors.append(f"stale language key in {language}: {old_key}")
        quatrefoil_keys = [key for key in data if "quatrefoil" in key and key.startswith("block.erydon.")]
        if len(quatrefoil_keys) < EXPECTED_QUATREFOIL_ALIASES:
            errors.append(f"{language} has only {len(quatrefoil_keys)} Quatrefoil block keys")

    stale_registry_references = 0
    old_registry_ids = {f"erydon:{entry.old_path}" for entry in entries}
    id_pattern = re.compile(r"erydon:[a-z0-9_]+")
    registry_reference_files = [
        *sorted((repo / "src/main/resources/data").rglob("*.json")),
        repo / "src/main/resources/assets/erydon/shaders/block.properties",
    ]
    for path in registry_reference_files:
        raw = path.read_text(encoding="utf-8-sig")
        stale = sorted(set(id_pattern.findall(raw)).intersection(old_registry_ids))
        if stale:
            stale_registry_references += sum(raw.count(value) for value in stale)
            errors.append(f"stale registry ID reference in {path}: {stale[:5]}")

    report: dict[str, object] = {
        "manifest_entries": len(entries),
        "permanent_aliases": sum(entry.permanent_alias for entry in entries),
        "direct_renames": sum(entry.mode == "DIRECT_RENAME" for entry in entries),
        "quatrefoil_aliases": sum(entry.permanent_alias and "quatrefoil" in entry.old_path
                                  for entry in entries),
        "canonical_registry_resources": canonical_resources,
        "stale_registry_resources": stale_resources,
        "canonical_language_keys": language_keys,
        "stale_language_keys": stale_language_keys,
        "stale_registry_references": stale_registry_references,
        "errors": errors,
    }
    return report


def main() -> int:
    arguments = parser().parse_args()
    try:
        if arguments.command == "export":
            entries = workbook_entries(arguments.workbook.resolve(), arguments.sheet)
            write_manifest(entries, arguments.output.resolve())
            print(json.dumps({
                "output": str(arguments.output.resolve()),
                "entries": len(entries),
                "aliases": sum(entry.permanent_alias for entry in entries),
                "direct_renames": sum(entry.mode == "DIRECT_RENAME" for entry in entries),
                "quatrefoil_aliases": sum(entry.permanent_alias and "quatrefoil" in entry.old_path
                                          for entry in entries),
            }, indent=2, ensure_ascii=False))
            return 0

        entries = read_manifest(arguments.manifest.resolve())
        if arguments.command == "apply":
            print(json.dumps(apply_source(arguments.repo, entries), indent=2))
            return 0

        report = audit_source(arguments.repo, entries)
        if arguments.report:
            arguments.report.parent.mkdir(parents=True, exist_ok=True)
            arguments.report.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n",
                                        encoding="utf-8")
        print(json.dumps(report, indent=2, ensure_ascii=False))
        return 1 if report["errors"] else 0
    except (OSError, ValueError, RuntimeError) as exception:
        print(f"id migration error: {exception}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
